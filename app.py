import streamlit as st
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re

# ── Page config ───────────────────────────────────────────────
st.set_page_config(
    page_title="PDF to Excel",
    page_icon="📄",
    layout="centered"
)

# ── Custom CSS ────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@400;600;700&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Sora', sans-serif;
}

.stApp {
    background: #0f0f13;
    color: #e8e6e0;
}

/* Header */
.hero {
    text-align: center;
    padding: 48px 0 32px;
}
.hero h1 {
    font-size: 2.8rem;
    font-weight: 700;
    color: #f0ede6;
    letter-spacing: -1px;
    margin-bottom: 8px;
}
.hero h1 span {
    color: #00e5a0;
}
.hero p {
    color: #888;
    font-size: 1rem;
    margin: 0;
}

/* Upload box */
[data-testid="stFileUploader"] {
    background: #1a1a22;
    border: 1.5px dashed #2e2e3a;
    border-radius: 16px;
    padding: 24px;
    transition: border-color 0.2s;
}
[data-testid="stFileUploader"]:hover {
    border-color: #00e5a0;
}

/* Buttons */
.stButton > button {
    background: #00e5a0;
    color: #0f0f13;
    font-family: 'Sora', sans-serif;
    font-weight: 600;
    font-size: 0.95rem;
    border: none;
    border-radius: 10px;
    padding: 12px 28px;
    width: 100%;
    transition: all 0.2s;
}
.stButton > button:hover {
    background: #00ffb3;
    transform: translateY(-1px);
    box-shadow: 0 8px 24px rgba(0,229,160,0.25);
}

/* Download button */
[data-testid="stDownloadButton"] > button {
    background: #1a1a22;
    color: #00e5a0;
    font-family: 'Sora', sans-serif;
    font-weight: 600;
    border: 1.5px solid #00e5a0;
    border-radius: 10px;
    padding: 12px 28px;
    width: 100%;
    transition: all 0.2s;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #00e5a0;
    color: #0f0f13;
}

/* Dataframe */
[data-testid="stDataFrame"] {
    background: #1a1a22;
    border-radius: 12px;
    border: 1px solid #2e2e3a;
}

/* Stats cards */
.stat-card {
    background: #1a1a22;
    border: 1px solid #2e2e3a;
    border-radius: 12px;
    padding: 20px;
    text-align: center;
}
.stat-num {
    font-size: 2rem;
    font-weight: 700;
    color: #00e5a0;
    font-family: 'DM Mono', monospace;
}
.stat-label {
    font-size: 0.8rem;
    color: #666;
    margin-top: 4px;
    text-transform: uppercase;
    letter-spacing: 1px;
}

/* Status badge */
.badge {
    display: inline-block;
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 0.78rem;
    font-weight: 600;
}
.badge-text { background: #0d2e1f; color: #00e5a0; }
.badge-ocr  { background: #2e1a0d; color: #ff9f43; }

/* Section label */
.section-label {
    font-size: 0.75rem;
    font-weight: 600;
    color: #555;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    margin-bottom: 12px;
}

/* Divider */
hr { border-color: #2e2e3a; }

/* Hide streamlit branding */
#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ── OCR check ─────────────────────────────────────────────────
def is_ocr_available():
    try:
        import pytesseract
        from pdf2image import convert_from_bytes
        return True
    except:
        return False

OCR_AVAILABLE = is_ocr_available()


# ── PDF text extraction ───────────────────────────────────────
def extract_text(file_bytes, filename):
    """Extract text — auto detects if OCR needed."""
    text = ""
    method = "text"

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            text += t

    # If very little text extracted → likely scanned → try OCR
    if len(text.strip()) < 50:
        if OCR_AVAILABLE:
            import pytesseract
            from pdf2image import convert_from_bytes
            method = "ocr"
            text = ""
            images = convert_from_bytes(file_bytes, dpi=300)
            for img in images:
                text += pytesseract.image_to_string(img)
        else:
            method = "ocr_unavailable"

    return text, method


# ── Table detection ───────────────────────────────────────────
def extract_tables_from_pdf(file_bytes):
    """Try to extract tables directly using pdfplumber."""
    all_tables = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            for tbl in tables:
                if tbl and len(tbl) > 1:
                    df = pd.DataFrame(tbl[1:], columns=tbl[0])
                    df = df.dropna(how='all').reset_index(drop=True)
                    all_tables.append({
                        'page': page_num,
                        'df': df
                    })
    return all_tables


# ── Text → rows (generic) ─────────────────────────────────────
def text_to_rows(text):
    """Convert extracted text into rows — generic approach."""
    rows = []
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    for line in lines:
        # Split by multiple spaces or tabs
        parts = re.split(r'\s{2,}|\t', line)
        if len(parts) > 1:
            rows.append(parts)
        else:
            rows.append([line])
    return rows


# ── Build styled Excel ────────────────────────────────────────
def build_excel(tables_data, text_rows, filename):
    wb = openpyxl.Workbook()

    header_fill  = PatternFill("solid", fgColor="00E5A0")
    header_font  = Font(name="Calibri", bold=True, color="0F0F13", size=11)
    alt_fill     = PatternFill("solid", fgColor="F2FBF7")
    white_fill   = PatternFill("solid", fgColor="FFFFFF")
    data_font    = Font(name="Calibri", size=10)
    bold_font    = Font(name="Calibri", bold=True, size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="D0D0D0")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    if tables_data:
        # ── Structured tables sheet ─────────────────────────
        for idx, tbl in enumerate(tables_data):
            sheet_name = f"Page{tbl['page']}_Table{idx+1}"[:31]
            ws = wb.create_sheet(sheet_name)
            df = tbl['df']

            # Headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(col_name) if col_name else f"Col{col_idx}")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

            # Data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                fill = alt_fill if row_idx % 2 == 0 else white_fill
                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
                    cell.font = data_font
                    cell.fill = fill
                    cell.alignment = left
                    cell.border = border

            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        # Remove default blank sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    else:
        # ── Raw text sheet ──────────────────────────────────
        ws = wb.active
        ws.title = "Extracted Text"

        ws.cell(row=1, column=1, value="Line No").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center
        ws.cell(row=1, column=1).border = border

        ws.cell(row=1, column=2, value="Content").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).alignment = center
        ws.cell(row=1, column=2).border = border

        for row_idx, row_data in enumerate(text_rows, 2):
            fill = alt_fill if row_idx % 2 == 0 else white_fill
            ws.cell(row=row_idx, column=1, value=row_idx - 1).font  = data_font
            ws.cell(row=row_idx, column=1).fill      = fill
            ws.cell(row=row_idx, column=1).alignment = center
            ws.cell(row=row_idx, column=1).border    = border

            content = "  |  ".join(row_data)
            ws.cell(row=row_idx, column=2, value=content).font  = data_font
            ws.cell(row=row_idx, column=2).fill      = fill
            ws.cell(row=row_idx, column=2).alignment = left
            ws.cell(row=row_idx, column=2).border    = border

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 80
        ws.freeze_panes = "A2"

    # ── Summary sheet ───────────────────────────────────────
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 40

    summary_data = [
        ("Source file",     filename),
        ("Tables found",    str(len(tables_data))),
        ("Sheets created",  str(len(wb.sheetnames) - 1)),
        ("Extracted by",    "PDF to Excel App"),
    ]

    ws_sum.cell(row=1, column=1, value="PDF to Excel — Summary").font = Font(name="Calibri", bold=True, size=13, color="0F0F13")
    ws_sum.cell(row=1, column=1).fill = header_fill
    ws_sum.merge_cells("A1:B1")
    ws_sum.cell(row=1, column=1).alignment = center
    ws_sum.row_dimensions[1].height = 30

    for row_idx, (key, val) in enumerate(summary_data, 3):
        ws_sum.cell(row=row_idx, column=1, value=key).font  = bold_font
        ws_sum.cell(row=row_idx, column=1).fill             = alt_fill
        ws_sum.cell(row=row_idx, column=1).alignment        = left
        ws_sum.cell(row=row_idx, column=1).border           = border
        ws_sum.cell(row=row_idx, column=2, value=val).font  = data_font
        ws_sum.cell(row=row_idx, column=2).fill             = white_fill
        ws_sum.cell(row=row_idx, column=2).alignment        = left
        ws_sum.cell(row=row_idx, column=2).border           = border

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ═══════════════════════════════════════════════════
#  UI
# ═══════════════════════════════════════════════════

st.markdown("""
<div class="hero">
    <h1>PDF to <span>Excel</span></h1>
    <p>Upload any PDF — tables extracted instantly, download as .xlsx</p>
</div>
""", unsafe_allow_html=True)

# OCR warning
if not OCR_AVAILABLE:
    st.warning("⚠️ OCR not installed — scanned/image PDFs may not extract well. Install with: `pip install pytesseract pdf2image` and `apt install tesseract-ocr poppler-utils`", icon="⚠️")

st.markdown('<div class="section-label">Upload your PDF</div>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

if uploaded_files:
    st.markdown("---")

    all_results = []

    for uploaded_file in uploaded_files:
        file_bytes = uploaded_file.read()
        filename   = uploaded_file.name

        with st.spinner(f"Processing **{filename}**..."):
            # Extract text
            text, method = extract_text(file_bytes, filename)

            # Try structured table extraction
            tables = extract_tables_from_pdf(file_bytes)

            # Fallback to text rows
            text_rows = text_to_rows(text) if not tables else []

            # Build Excel
            excel_bytes = build_excel(tables, text_rows, filename)

        all_results.append({
            'filename': filename,
            'tables':   tables,
            'text_rows': text_rows,
            'text':     text,
            'method':   method,
            'excel':    excel_bytes
        })

    # ── Results ──────────────────────────────────────────────
    for result in all_results:
        st.markdown(f"### 📄 {result['filename']}")

        # Method badge
        if result['method'] == 'text':
            st.markdown('<span class="badge badge-text">✓ Text PDF — direct extraction</span>', unsafe_allow_html=True)
        elif result['method'] == 'ocr':
            st.markdown('<span class="badge badge-ocr">⚡ Scanned PDF — OCR used</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="badge badge-ocr">⚠ Scanned PDF — install OCR for better results</span>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Stats
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f'<div class="stat-card"><div class="stat-num">{len(result["tables"])}</div><div class="stat-label">Tables found</div></div>', unsafe_allow_html=True)
        with col2:
            rows = sum(len(t['df']) for t in result['tables']) if result['tables'] else len(result['text_rows'])
            st.markdown(f'<div class="stat-card"><div class="stat-num">{rows}</div><div class="stat-label">Rows extracted</div></div>', unsafe_allow_html=True)
        with col3:
            size_kb = round(len(result['excel']) / 1024, 1)
            st.markdown(f'<div class="stat-card"><div class="stat-num">{size_kb}K</div><div class="stat-label">Excel file size</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Preview
        if result['tables']:
            st.markdown('<div class="section-label">Preview — first table</div>', unsafe_allow_html=True)
            preview_df = result['tables'][0]['df'].head(10)
            st.dataframe(preview_df, use_container_width=True)
        elif result['text']:
            st.markdown('<div class="section-label">Preview — extracted text</div>', unsafe_allow_html=True)
            preview_lines = result['text'][:800] + ("..." if len(result['text']) > 800 else "")
            st.text_area("", value=preview_lines, height=180, label_visibility="collapsed")

        # Download
        out_name = result['filename'].replace('.pdf', '.xlsx')
        st.download_button(
            label=f"⬇ Download {out_name}",
            data=result['excel'],
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{result['filename']}"
        )

        st.markdown("---")

else:
    # Empty state
    st.markdown("""
    <div style="text-align:center; padding: 48px 0; color: #444;">
        <div style="font-size: 3rem; margin-bottom: 16px;">📂</div>
        <div style="font-size: 1rem;">Drop your PDF above to get started</div>
        <div style="font-size: 0.85rem; margin-top: 8px; color: #333;">Supports text PDFs and scanned documents</div>
    </div>
    """, unsafe_allow_html=True)

# Footer
st.markdown("""
<div style="text-align:center; padding: 32px 0 16px; color: #333; font-size: 0.78rem;">
    PDF to Excel App &nbsp;·&nbsp; Built with Streamlit &nbsp;·&nbsp; Free to use
</div>
""", unsafe_allow_html=True)
